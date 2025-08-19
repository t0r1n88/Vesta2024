"""
Скрипт для пакетного создания документов из шаблонов
"""
import pandas as pd
import openpyxl
from tkinter import messagebox
import os
import re
import datetime
from docxtpl import DocxTemplate
from docx.opc.exceptions import PackageNotFoundError
import tempfile
from docxcompose.composer import Composer
from docx import Document
from tkinter import messagebox
from jinja2 import exceptions
import time
import warnings
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
warnings.simplefilter(action='ignore', category=DeprecationWarning)
warnings.simplefilter(action='ignore', category=FutureWarning)
warnings.simplefilter(action='ignore', category=UserWarning)
pd.options.mode.chained_assignment = None


class SameNameColumn(Exception):
    """
    Исключение для обработки случая когда в двух листах есть одинаковые названия колонок
    """
    pass

class SamePathFolder(Exception):
    """
    Исключение для случая когда одна и та же папка выбрана в качестве источника и конечной папки
    """
    pass

class NotReqSheet(Exception):
    """
    Исключение для проверки наличия минимум 2 листов
    """
    pass

class NotFileSource(Exception):
    """
    Исключение для обработки случая когда не найдены файлы внутри исходной папки
    """
    pass

class NotColumn(Exception):
    """
    Исключение для обработки случая когда не найдена колонка по которой будут делаться названия файлов
    """
    pass


def create_doc_convert_date(cell):
    """
    Функция для конвертации даты при создании документов
    :param cell:
    :return:
    """
    try:
        string_date = datetime.datetime.strftime(cell, '%d.%m.%Y')
        return string_date
    except ValueError:
        return ''
    except TypeError:
        return ''
def convert_string_date(df:pd.DataFrame,lst_date_columns:list)->pd.DataFrame:
    """
    Функция для коневертации колонок с датами в строковый формат для правильного отображения
    :param df: датафрейм с данными
    :param lst_date_columns: список с индексами колонок с датами
    :return: исправленный датафрейм
    """
    lst_name_columns = [] # список куда будут сохраняться названия колонок
    for i in lst_date_columns:
        lst_name_columns.append(list(df.columns)[i])

    # Конвертируем в пригодный строковый формат
    for name_column in lst_name_columns:
        df[name_column] = pd.to_datetime(df[name_column],errors='ignore')
        df[name_column] = df[name_column].apply(create_doc_convert_date)


    return df

def convert_to_date(value):
    """
    Функция для конвертации строки в текст
    :param value: значение для конвертации
    :return:
    """
    try:
        date_value = datetime.datetime.strptime(value, '%Y-%m-%d %H:%M:%S')
        return date_value
    except ValueError:
        # стандартный формат
        result = re.search(r'^\d{2}\.\d{2}\.\d{4}$', value)
        if result:
            return datetime.datetime.strptime(result.group(0), '%d.%m.%Y')
        # формат яндекс форм
        second_result = re.search(r'^\d{4}-\d{2}-\d{2}$', value)
        if second_result:
            return datetime.datetime.strptime(second_result.group(0), '%Y-%m-%d')
        else:
            return ''
    except:
        return ''

def selection_name_column(lst_cols: list, pattern: str):
    """
    Функция для отбора значений попадающих под условие
    :param lst_cols: список с строками
    :param pattern: паттерн отбора
    :return:кортеж из 2 списков, первй список это подошедшие под условие а второй список это не подошедшие
    """
    valid_cols = [name_col for name_col in lst_cols if re.search(pattern,name_col)]
    not_valid_cols = (set(lst_cols)).difference(set(valid_cols))
    return valid_cols,not_valid_cols


def copy_folder_structure(source_folder:str,destination_folder:str):
    """
    Функция для копирования структуры папок внутри выбраной папки
    :param source_folder: Исходная папка
    :param destination_folder: конечная папка
    :return: Структура папок как в исходной папке
    """
    # Получаем список папок внутри source_folder

    lst_subdirs =  [] # список для подпапок
    lst_files = [] # список для файлов
    lst_source_folders = [] # список для хранения путей к папкам в исходной папке

    for dirname, dirnames, filenames in os.walk(source_folder):
        # print path to all subdirectories first.
        for subdirname in dirnames:
            lst_subdirs.append(subdirname)
            lst_source_folders.append(f'{dirname}/{subdirname}')

    # ищем файлы
    for dirname, dirnames, filenames in os.walk(source_folder):
        for file in filenames:
            lst_files.append(file)

    # заменяем папку назначения
    lst_dest_folders = [path.replace(source_folder,destination_folder) for path in lst_source_folders]
    for path_folder in lst_dest_folders:
        if not os.path.exists(path_folder):
            os.makedirs(path_folder)
    # создаем словарь где ключ это путь к папкам в исходном файле а значение это путь к папкам в конечной папке
    # проверяем количество найденных папок
    if len(lst_subdirs) != 0:
        dct_path = dict(zip(lst_source_folders,lst_dest_folders))
    else:
        # если подпапок нет то сохраняем в итоговую папку
        dct_path = {source_folder:destination_folder}

    if len(lst_files) == 0:
        raise NotFileSource

    return dct_path

def combine_all_docx(filename_master, files_lst,path_to_end_folder_doc,name_template:str):
    """
    Функция для объединения файлов Word взято отсюда
    https://stackoverflow.com/questions/24872527/combine-word-document-using-python-docx
    :param filename_master: базовый файл
    :param files_list: список с созданными файлами
    :return: итоговый файл
    """

    # Получаем текущее время
    t = time.localtime()
    current_time = time.strftime('%H_%M_%S', t)

    number_of_sections = len(files_lst)
    # Открываем и обрабатываем базовый файл
    master = Document(filename_master)
    composer = Composer(master)
    # Перебираем и добавляем файлы к базовому
    for i in range(0, number_of_sections):
        doc_temp = Document(files_lst[i])
        composer.append(doc_temp)
    # Сохраняем файл
    composer.save(f"{path_to_end_folder_doc}/{name_template}_ОБЩИЙ файл от {current_time}.docx")




def generate_docs(dct_descr:dict,data_df:pd.DataFrame,source_folder:str,destination_folder:str,name_column:str):
    """
    Основная функция генерации документов
    :param dct_descr: словарь с константами
    :param data_df: датафрейм с изменяющимися данными
    :param source_folder: исходная папка
    :param destination_folder: конечная папка
    :param name_column: название колонки по которой будут называться раздельные файлы
    :return: Пакет документации в формате docx
    """

    # добавляем колонки из описания программы в датафрейм с общими данными
    for key, value in dct_descr.items():
        data_df[key] = value

    lst_data_df = data_df.copy()  # копируем датафрейм

    # Конвертируем датафрейм в список словарей
    data = data_df.to_dict('records')
    dct_path = copy_folder_structure(source_folder, destination_folder)  # копируем структуру папок

    # начинаем обработку папок и файлов внутри них
    for source_folder, dest_folder in dct_path.items():
        for file in os.listdir(source_folder):
            if file.endswith('.docx') and not file.startswith('~$'):  # получаем только файлы docx и не временные
                name_template = source_folder.split('/')[-1] # получаем название шаблона
                name_template = re.sub(r'[\r\b\n\t<>:"?*|\\/]', '_', name_template)
                # определяем тип создаваемого документа
                if 'раздельный' in file.lower():
                    used_name_file = set()  # множество для уже использованных имен файлов
                    # Создаем в цикле документы
                    for idx, row in enumerate(data):
                        doc = DocxTemplate(f'{source_folder}/{file}')
                        context = row
                        doc.render(context)
                        # Сохраняенм файл
                        # получаем название файла и убираем недопустимые символы < > : " /\ | ? *
                        name_file = row[name_column]
                        name_file = re.sub(r'[\r\b\n\t<>:"?*|\\/]', '_', str(name_file))

                        if name_file[:80] in used_name_file:
                            name_file = f'{name_file[:75]}_{idx}'

                        doc.save(f'{dest_folder}/{name_template}_{name_file[:80]}.docx')
                        used_name_file.add(name_file[:80])  # добавляем в использованные названия
                elif 'общий' in file.lower():
                    # Список с созданными файлами
                    files_lst = []
                    # Создаем временную папку
                    with tempfile.TemporaryDirectory() as tmpdirname:
                        print('created temporary directory', tmpdirname)
                        # Создаем и сохраняем во временную папку созданные документы Word
                        for idx, row in enumerate(data):
                            doc = DocxTemplate(f'{source_folder}/{file}')
                            context = row
                            doc.render(context)
                            # Сохраняем файл
                            # очищаем от запрещенных символов
                            name_file = row[name_column]

                            name_file = re.sub(r'[\r\b\n\t<> :"?*|\\/]', '_', name_file)

                            doc.save(f'{tmpdirname}/{name_file[:80]}_{idx}.docx')
                            # Добавляем путь к файлу в список
                            files_lst.append(f'{tmpdirname}/{name_file[:80]}_{idx}.docx')
                        # Получаем базовый файл
                        if len(files_lst) != 0:  # проверка на заполнение листа с данными
                            main_doc = files_lst.pop(0)
                            # Запускаем функцию
                            combine_all_docx(main_doc, files_lst, dest_folder,name_template)
                else:
                    # генерируем текущее время
                    t = time.localtime()
                    current_time = time.strftime('%H_%M_%S', t)
                    used_name_file = set()  # множество для уже использованных имен файлов
                    doc = DocxTemplate(f'{source_folder}/{file}')
                    context = dict()
                    context['Итог'] = lst_data_df.to_dict('records')
                    context.update(dct_descr)  # добавляем словарь с описанием программы

                    doc.render(context)
                    # Сохраняем файл
                    # получаем название файла и убираем недопустимые символы < > : " /\ | ? *
                    name_file = file.split('.docx')[0]
                    name_file = re.sub('Шаблон ', '', name_file)
                    name_file = re.sub(r'[\r\b\n\t<>:"?*|\\/]', '_', name_file)

                    # проверяем файл на наличие, если файл с таким названием уже существует то добавляем окончание
                    if name_file[:80] in used_name_file:
                        name_file = f'{name_file[:75]}_{idx}'

                    doc.save(f'{dest_folder}/{name_file[:80]} {current_time}.docx')
                    used_name_file.add(name_file[:80])








def processing_create_batch_documents(data_file:str,folder_template:str,result_folder:str,name_file_column:str):
    """
    Скрипт для пакетного создания документов. Точка входа
    :param data_file: файл Excel с данными
    :param folder_template: папка с шаблонами
    :param result_folder: итоговая папка
    :param name_file_column: название колонки по которой будут создаваться названия файлов
    """
    try:
        if folder_template == result_folder:
            raise SamePathFolder

        req_wb = openpyxl.load_workbook(data_file)  # загружаем файл для подсчета количества листов
        if len(req_wb.sheetnames) < 2:
            raise NotReqSheet

        descr_sheet = req_wb.sheetnames[0] # название листа с константами
        data_sheet = req_wb.sheetnames[1] # названия листа с данными
        # Обрабатываем лист с константами
        descr_df = pd.read_excel(data_file, sheet_name=descr_sheet, dtype=str, usecols='A:B')  # получаем данные констант
        descr_df.dropna(how='all', inplace=True)  # удаляем пустые строки

        # Предобработка датафрейма с данными слушателей
        data_df = pd.read_excel(data_file, sheet_name=data_sheet, dtype=str)  # получаем данные
        # Проверяем наличие нужных колонок в файле с данными
        data_df.dropna(how='all', inplace=True)  # удаляем пустые строки
        if name_file_column not in data_df.columns:
            raise NotColumn

        req_wb.close() # закрываем файл

        descr_df = descr_df.transpose()
        descr_df.columns = descr_df.iloc[0]  # устанавливаем первую строку в качестве названий колонок
        descr_df = descr_df.iloc[1:]  # удаляем первую строку
        descr_df.index = [0] # переименовываем оставшийся индекс в 0
        descr_df = descr_df.applymap(
            lambda x: re.sub(r'\s+', ' ', x) if isinstance(x, str) else x)  # очищаем от лишних пробелов
        descr_df = descr_df.applymap(
            lambda x: x.strip() if isinstance(x, str) else x)  # очищаем от пробелов в начале и конце

        # делаем строковыми названия колонок
        descr_df.columns = list(map(str,descr_df.columns))
        data_df.columns = list(map(str,data_df.columns))

        # проверяем на совпадение названий колонок в обоих листах
        intersection_columns = set(descr_df.columns).intersection(set(data_df.columns))
        if len(intersection_columns) > 0:
            raise SameNameColumn

        # Обрабатываем колонки с датами в описании
        lst_date_columns_descr = []  # список для колонок с датами
        for idx, column in enumerate(descr_df.columns):
            if 'дата' in column.lower():
                lst_date_columns_descr.append(idx)

        descr_df = convert_string_date(descr_df,lst_date_columns_descr)
        # обрабатываем колонки с датами в списке
        lst_date_columns_data = []  # список для колонок с датами
        for idx, column in enumerate(data_df.columns):
            if 'дата' in column.lower():
                lst_date_columns_data.append(column)
        data_df[lst_date_columns_data] = data_df[lst_date_columns_data].applymap(convert_to_date)  # Приводим к типу дата


        # приводим даты к строковому типу
        data_df[lst_date_columns_data] = data_df[lst_date_columns_data].applymap(
            lambda x: x.strftime('%d.%m.%Y') if isinstance(x, (pd.Timestamp, datetime.datetime)) and pd.notna(x) else x)

        # получаем списки валидных названий колонок
        descr_valid_cols,descr_not_valid_cols = selection_name_column(list(descr_df.columns),r'^[a-zA-ZЁёа-яА-Я_]+$')
        data_valid_cols, data_not_valid_cols = selection_name_column(list(data_df.columns),r'^[a-zA-ZЁёа-яА-Я_]+$')

        data_df[name_file_column] = data_df[name_file_column].fillna('_')
        # заполняем наны пробелами
        descr_df.fillna(' ',inplace=True)
        data_df.fillna(' ',inplace=True)

        # Словарь с константами
        dct_descr = dict()
        for name_column in descr_valid_cols:
            dct_descr[name_column] = descr_df.loc[0,name_column]


        generate_docs(dct_descr,data_df[data_valid_cols],folder_template,result_folder,name_file_column)
    except FileNotFoundError as e:
        messagebox.showerror('Веста Обработка таблиц и создание документов',
                             f'Не удалось создать файл с названием {e}\n'
                             f'Уменьшите количество символов в соответствующей строке файла с данными в колонке по которой создаются имена файлов или выберите более короткий путь к итоговой папке')
    except PackageNotFoundError as e:
        messagebox.showerror('Веста Обработка таблиц и создание документов',
                             f'Не удалось создать файл с названием {e}\n'
                             f'Уменьшите количество символов в соответствующей строке файла с данными в колонке по которой создаются имена файлов или выберите более короткий путь к итоговой папке')

    except NotFileSource:
        messagebox.showerror('Веста Обработка таблиц и создание документов',
                             f'В папке с шаблонами не найдены файлы docx !!!')
    except exceptions.TemplateSyntaxError:
        messagebox.showerror('Веста Обработка таблиц и создание документов',
                             f'Ошибка в оформлении вставляемых значений в шаблоне\n'
                             f'Проверьте свой шаблон на наличие следующих ошибок:\n'
                             f'1) Вставляемые значения должны быть оформлены двойными фигурными скобками\n'
                             f'{{{{Вставляемое_значение}}}}\n'
                             f'2) В названии колонки в таблице откуда берутся данные - есть пробелы,цифры,знаки пунктуации и т.п.\n'
                             f'в названии колонки должны быть только буквы и нижнее подчеркивание.\n'
                             f'{{{{Дата_рождения}}}}')
    except NotReqSheet:
        messagebox.showerror('Веста Обработка таблиц и создание документов',
                             f'В файле с данными должно быть минимум 2 листа')
    except NotColumn:
        messagebox.showerror('Веста Обработка таблиц и создание документов',
                             f'На листе с изменяемыми данными (т.е. второй лист по порядку) отсутствует колонка по значениям которой будут создаваться названия файлов')

    else:
        messagebox.showinfo('Веста Обработка таблиц и создание документов', 'Создание документов успешно завершено !')






























if __name__ == '__main__':
    main_data_file = 'c:/Users/1/PycharmProjects/Vesta2024/data/Пакетное создание документов/Данные для заполнения ХАССП.xlsx'
    main_folder_template = 'c:/Users/1/PycharmProjects/Vesta2024/data/Пакетное создание документов/Формы контроля'
    main_result_folder = 'c:/Users/1/PycharmProjects/Vesta2024/data/Результат'
    main_name_column = 'Наименование_юрлица'

    processing_create_batch_documents(main_data_file,main_folder_template,main_result_folder,main_name_column)

    print('Lindy Booth')

