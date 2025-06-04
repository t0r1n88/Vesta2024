"""
Извлечение данных из файлов Excel со сложной структурой
"""
import os
import pandas as pd
from tkinter import messagebox
import openpyxl
import time
import re
import warnings
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
warnings.simplefilter(action='ignore', category=DeprecationWarning)
warnings.simplefilter(action='ignore', category=FutureWarning)
warnings.simplefilter(action='ignore', category=UserWarning)
pd.options.mode.chained_assignment = None
import logging
logging.basicConfig(
    level=logging.WARNING,
    filename="error.log",
    filemode='w',
    # чтобы файл лога перезаписывался  при каждом запуске.Чтобы избежать больших простыней. По умолчанию идет 'a'
    format="%(asctime)s - %(module)s - %(levelname)s - %(funcName)s: %(lineno)d - %(message)s",
    datefmt='%H:%M:%S',)

class NotFile(Exception):
    """
    Обработка случаев когда нет файлов в папке
    """
    pass

class NotCorrectParams(Exception):
    """
    Исключение для случаев когда нет ни одного корректного параметра
    """
    pass


def count_text_value(df):
    """
    Функция для подсчета количества вариантов того или иного показателя
    :param df: датафрейм с сырыми данными. Название показателя значение показателя(строка разделенная ;)
    :return: обработанный датафрейм с мультиндексом, где (Название показателя это индекс верхнего уровня, вариант показателя это индекс второго уровня а значение это сколько раз встречался
    этот вариант в обрабатываемых файлах)
    """
    data = dict()

    #
    for row in df.itertuples():
        value = row[2]
        if type(value) == float or type(value) == int:
            continue
        # Создаем список, разделяя строку по ;
        lst_value = row[2].split(';')[:-1]
        #     # Отрезаем последний элемент, поскольку это пустое значение
        temp_df = pd.DataFrame({'Value': lst_value})
        counts_series = temp_df['Value'].value_counts()
        # Делаем индекс колонкой и превращаем в обычную таблицу
        index_count_values = counts_series.reset_index()
        # Итерируемся по таблице.Это делается чтобы заполнить словарь на основе которого будет создаваться итоговая таблица
        for count_row in index_count_values.itertuples():
            # Заполняем словарь
            data[(row[1], count_row[1])] = count_row[2]
    # Создаем на основе получившегося словаря таблицу
    out_df = pd.Series(data).to_frame().reset_index()
    out_df = out_df.set_index(['level_0', 'level_1'])
    out_df.index.names = ['Название показателя', 'Вариант показателя']
    out_df.rename(columns={0: 'Количество'}, inplace=True)
    return out_df


def check_data(cell, text_mode):
    """
    Функция для проверки значения ячейки. Для обработки пустых значений, строковых значений, дат
    :param cell: значение ячейки
    :return: 0 если значение ячейки не число
            число если значение ячейки число(ха звучит глуповато)
    думаю функция должна работать с дополнительным параметром, от которого будет зависеть подсчет значений навроде галочек или плюсов в анкетах или опросах.
    """
    # Проверяем режим работы. если текстовый, то просто складываем строки
    if text_mode == 'Yes':
        if cell is None:
            return ''
        else:
            temp_str = str(cell)
            return f'{temp_str};'
    # Если режим работы стандартный. Убрал подсчет строк и символов в числовом режиме, чтобы не запутывать.
    else:
        if cell is None:
            return 0
        if type(cell) == int:
            return cell
        elif type(cell) == float:
            return cell
        else:
            return 0

def count_files_with_extension(folder_path, extension):
    # Получаем список файлов в указанной папке
    files = os.listdir(folder_path)

    # Считаем количество файлов с заданным разрешением
    count = sum(1 for file in files if not file.startswith('~$') and file.endswith(extension))

    return count

def check_range(df:pd.DataFrame):
    """
    Функция для проверки корректности записи диапазона
    :param датафрейм: параметры
    :return: Правильно или Неправильно
    """
    error_df = pd.DataFrame(
        columns=['Название файла','Описание ошибки'])  # датафрейм для ошибок

    pattern = r'^([A-Z]{1,3})(\d+)'

    out_dct = {} # словарь для хранения параметров
    # перебираем строки
    for row in df.itertuples():
        address_cell = row[2] # адрес ячейки
        # Проверяем на корректность диапазон
        result = re.search(pattern,address_cell)
        if result:
            out_dct[f'{row[1]}'] = f'{row[2]}'

        else:
            temp_error_df = pd.DataFrame(columns=['Название файла', 'Описание ошибки'],
                data=[['Ошибка в файле параметров', f'Ошибка в написании адреса ячейки- {address_cell} . Правильный формат написания 1,2,3 буквы латинского алфавита потом число. Например AZ10']])
            error_df = pd.concat([error_df,temp_error_df])
    return out_dct,error_df





def extract_data_from_hard_xlsx(mode_text, name_file_params_calculate_data, files_calculate_data, path_to_end_folder_calculate_data):
    """
    Функция для извлечения данных из таблиц Excel со сложной структурой, извлечение происходит из конкретных ячеек указанных в файле параметров
    :param mode_text: режим работы (обработка текста или чисел)
    :param name_file_params_calculate_data: файл  указанием ячеек данные из которых нужно извлечь
    :param files_calculate_data:папка с файлами которые нужно обработать
    :param path_to_end_folder_calculate_data:  итоговая папка
    :return:
    """
    try:
        count = 0
        count_errors = 0
        lst_files = [] # список для файлов
        for dirpath, dirnames, filenames in os.walk(files_calculate_data):
            lst_files.extend(filenames)
        # отбираем файлы
        lst_xlsx = [file for file in lst_files if not file.startswith('~$') and file.endswith('.xlsx')]

        quantity_files = len(lst_xlsx)  # считаем сколько xlsx файлов в папке
        current_time = time.strftime('%H_%M_%S')

        error_df = pd.DataFrame(
            columns=['Название файла','Описание ошибки'])  # датафрейм для ошибок

        # Обрабатываем в зависимости от количества файлов в папке
        if quantity_files == 0:
            raise NotFile
        elif quantity_files == 1:
            try:

                # Получаем шаблон с данными, первую строку пропускаем, поскольку название обрабатываемого листа мы уже получили

                params_df = pd.read_excel(name_file_params_calculate_data)
                params_df.dropna(how='any', inplace=True)  # очищаем от неполных строк
                params_df = params_df.astype(str) # приводим к строковому виду

            except NameError:
                messagebox.showerror('Эльпида Школьная отчетность',
                                     f'Выберите файл с параметрами обработки')
            except:
                messagebox.showerror('Эльпида Школьная отчетность',
                                     f'Не удалось обработать файл с параметрами. Возможно файл поврежден')

            # Создаем словарь параметров
            param_dict, error_params_df = check_range(
                params_df)  # создаем словарь с параметрами и датафрейм с ошибками

            # добавляем в датафрейм ошибок
            error_df = pd.concat([error_df, error_params_df], ignore_index=True)
            if len(param_dict) == 0:
                raise NotCorrectParams


            if mode_text == 'Yes':
                result_dct = {key: '' for key, value in param_dict.items()}
            else:
                result_dct = {key: 0 for key, value in param_dict.items()}

            lst_cols = ['Название листа'].extend((list(param_dict.keys())))
            check_df = pd.DataFrame(columns=lst_cols)
            for dirpath, dirnames, filenames in os.walk(files_calculate_data):
                for file in filenames:
                    if not file.startswith('~$') and file.endswith('.xlsx'):
                        name_file = file.split('.xlsx')[0]
                        print(name_file) # обрабатываемый файл

                        try:
                            wb = openpyxl.load_workbook(f'{dirpath}/{file}',data_only=True) # открываем файл
                        except:
                            temp_error_df = pd.DataFrame(
                                data=[[f'{file}',
                                       f'Не удалось обработать файл. Возможно файл поврежден'
                                       ]],
                                columns=['Название файла',
                                         'Описание ошибки'])
                            error_df = pd.concat([error_df, temp_error_df], axis=0,
                                                 ignore_index=True)
                            count_errors += 1
                            continue
                        # перебираем листы в файле
                        # Создаем словарь для создания строки которую мы будем добавлять в проверочный датафрейм
                        for name_list in wb.sheetnames:
                            new_row = dict()
                            new_row['Название листа'] = name_list  # создаем ключ по названию файла
                            sheet = wb[name_list]
                            for key, cell in param_dict.items():
                                try:
                                    result_dct[key] += check_data(sheet[cell].value, mode_text)  # извлекаем данные из ячейки
                                    new_row[key] = sheet[cell].value
                                except ValueError:
                                    temp_error_df = pd.DataFrame(
                                        data=[[f'{file}', f'При извлечении данных показателя {key} из ячейки {cell} возникла ошибка. Проверьте правильность написания адреса ячейки. Правильный адрес это A2,F1211'
                                               ]],
                                        columns=['Название файла',
                                                 'Описание ошибки'])
                                    error_df = pd.concat([error_df, temp_error_df], axis=0,
                                                         ignore_index=True)
                                    count_errors += 1
                                    continue
                                except:
                                    temp_error_df = pd.DataFrame(
                                        data=[[f'{file}',
                                               f'При извлечении данных показателя {key} из ячейки {cell} возникла неопределенная ошибка'
                                               ]],
                                        columns=['Название файла',
                                                 'Описание ошибки'])
                                    error_df = pd.concat([error_df, temp_error_df], axis=0,
                                                         ignore_index=True)
                                    count_errors += 1
                                    continue

                            temp_df = pd.DataFrame(new_row, index=['temp_index'])
                            check_df = pd.concat([check_df, temp_df], ignore_index=True)
                    count += 1 # считаем обработанные файлы
                    # сохраняем
                    error_df.to_excel(f'{path_to_end_folder_calculate_data}/Ошибки {current_time}.xlsx',
                                      index=False)
                    check_df.to_excel(
                        f'{path_to_end_folder_calculate_data}/Проверка вычисления {current_time}.xlsx', index=False)

                    # Создание итоговой таблицы результатов подсчета

                    finish_result = pd.DataFrame()

                    if count != 0:
                        finish_result['Наименование показателя'] = result_dct.keys()
                        finish_result['Значение показателя'] = result_dct.values()
                        # Проводим обработку в зависимости от значения переключателя
                        if mode_text == 'Yes':
                            # Обрабатываем датафрейм считая текстовые данные
                            count_text_df = count_text_value(finish_result)
                            # сохраняем
                            count_text_df.to_excel(
                                f'{path_to_end_folder_calculate_data}/Подсчет текстовых значений {current_time}.xlsx')
                        else:
                            # сохраняем

                            finish_result.to_excel(
                                f'{path_to_end_folder_calculate_data}/Итоговые значения {current_time}.xlsx',
                                index=False)

                        if count_errors != 0:
                            messagebox.showinfo('Эльпида Школьная отчетность',
                                                f'В некоторых файлах обнаружены ошибки!\nОбработано файлов:  {count} из {quantity_files}\n Необработанные файлы указаны в файле {path_to_end_folder_calculate_data}/Ошибки {current_time}.xlsx')
                        else:
                            messagebox.showinfo('Эльпида Школьная отчетность',
                                                f'Обработка файлов успешно завершена!\nОбработано файлов:  {count} из {quantity_files}')
                    else:
                        messagebox.showwarning('Эльпида Школьная отчетность',
                                               f'Обработано {count} из {quantity_files} файлов.\n Причины необработки файлов указаны в файле {path_to_end_folder_calculate_data}/Ошибки {current_time}.xlsx')


        else:

            try:
                name_list_df = pd.read_excel(name_file_params_calculate_data, nrows=1,usecols='A:B')
                name_list_df.columns = ['Показатель','Значение']
                name_list = name_list_df['Значение'].loc[0]

                # Получаем шаблон с данными, первую строку пропускаем, поскольку название обрабатываемого листа мы уже получили

                params_df = pd.read_excel(name_file_params_calculate_data, skiprows=1)
                params_df.dropna(how='any', inplace=True)  # очищаем от неполных строк
                params_df = params_df.astype(str) # приводим к строковому виду
            except NameError:
                messagebox.showerror('Эльпида Школьная отчетность',
                                     f'Выберите файл с параметрами обработки')
            except:
                messagebox.showerror('Эльпида Школьная отчетность',
                                     f'Не удалось обработать файл с параметрами. Возможно файл поврежден')


            # Создаем словарь параметров
            param_dict, error_params_df = check_range(
                params_df)  # создаем словарь с параметрами и датафрейм с ошибками

            # добавляем в датафрейм ошибок
            error_df = pd.concat([error_df, error_params_df], ignore_index=True)
            if len(param_dict) == 0:
                raise NotCorrectParams

            if mode_text == 'Yes':
                result_dct = {key: '' for key, value in param_dict.items()}
            else:
                result_dct = {key: 0 for key, value in param_dict.items()}

                # Создаем датафрейм для контроля процесса подсчета и заполняем словарь на основе которого будем делать итоговую таблицу
            check_df = pd.DataFrame(columns=list(param_dict.keys()))
            # Вставляем колонку для названия файла
            check_df.insert(0, 'Название файла', '')
            for dirpath, dirnames, filenames in os.walk(files_calculate_data):
                for file in filenames:
                    if not file.startswith('~$') and file.endswith('.xlsx'):
                        name_file = file.split('.xlsx')[0]
                        print(name_file) # обрабатываемый файл
                        # Проверяем чтобы файл не был резервной копией или файлом с другим расширением.
                        # Создаем словарь для создания строки которую мы будем добавлять в проверочный датафрейм
                        new_row = dict()
                        new_row['Название файла'] = name_file # создаем ключ по названию файла
                        try:
                            wb = openpyxl.load_workbook(f'{dirpath}/{file}',data_only=True) # открываем файл
                        except:
                            temp_error_df = pd.DataFrame(
                                data=[[f'{file}',
                                       f'Не удалось обработать файл. Возможно файл поврежден'
                                       ]],
                                columns=['Название файла',
                                         'Описание ошибки'])
                            error_df = pd.concat([error_df, temp_error_df], axis=0,
                                                 ignore_index=True)
                            count_errors += 1
                            continue
                        # Проверяем наличие листа
                        if name_list in wb.sheetnames:
                            sheet = wb[name_list]
                            # перебираем все указанные адреса ячеек
                            for key, cell in param_dict.items():
                                try:
                                    result_dct[key] += check_data(sheet[cell].value, mode_text)  # извлекаем данные из ячейки
                                    new_row[key] = sheet[cell].value
                                except ValueError:
                                    temp_error_df = pd.DataFrame(
                                        data=[[f'{file}', f'При извлечении данных показателя {key} из ячейки {cell} возникла ошибка. Проверьте правильность написания адреса ячейки. Правильный адрес это A2,F1211'
                                               ]],
                                        columns=['Название файла',
                                                 'Описание ошибки'])
                                    error_df = pd.concat([error_df, temp_error_df], axis=0,
                                                         ignore_index=True)
                                    count_errors += 1
                                    continue
                                except:
                                    temp_error_df = pd.DataFrame(
                                        data=[[f'{file}',
                                               f'При извлечении данных показателя {key} из ячейки {cell} возникла неопределенная ошибка'
                                               ]],
                                        columns=['Название файла',
                                                 'Описание ошибки'])
                                    error_df = pd.concat([error_df, temp_error_df], axis=0,
                                                         ignore_index=True)
                                    count_errors += 1
                                    continue

                            temp_df = pd.DataFrame(new_row, index=['temp_index'])
                            check_df = pd.concat([check_df, temp_df], ignore_index=True)
                            count += 1 # считаем обработанные файлы
                        else:
                            # Записываем ошибку
                            temp_error_df = pd.DataFrame(data=[[f'{file}', f'Среди листов {wb.sheetnames} не найден лист {name_list}. Название листа из которого вы хотите получить данные должно быть на 2 строке файла параметров в ячейке B2'
                                                                ]],
                                                         columns=['Название файла',
                                                                  'Описание ошибки'])
                            error_df = pd.concat([error_df, temp_error_df], axis=0,
                                                           ignore_index=True)
                            count_errors += 1
                            continue
                    else:
                        if file.startswith('~$'):
                            continue
                        else:
                            # Записываем ошибку
                            temp_error_df = pd.DataFrame(
                                data=[[f'{file}', 'Неверное расширение файла! Обрабатываются только файлы с расширением XLSX!'
                                       ]],
                                columns=['Название файла',
                                         'Описание ошибки'])
                            error_df = pd.concat([error_df, temp_error_df], axis=0,
                                                 ignore_index=True)
                            count_errors += 1


            # сохраняем
            error_df.to_excel(f'{path_to_end_folder_calculate_data}/Ошибки {current_time}.xlsx',index=False)
            check_df.to_excel(f'{path_to_end_folder_calculate_data}/Проверка вычисления {current_time}.xlsx', index=False)

            # Создание итоговой таблицы результатов подсчета

            finish_result = pd.DataFrame()

            if count !=0:
                finish_result['Наименование показателя'] = result_dct.keys()
                finish_result['Значение показателя'] = result_dct.values()
                # Проводим обработку в зависимости от значения переключателя
                if mode_text == 'Yes':
                    # Обрабатываем датафрейм считая текстовые данные
                    count_text_df = count_text_value(finish_result)
                    # сохраняем
                    count_text_df.to_excel(
                        f'{path_to_end_folder_calculate_data}/Подсчет текстовых значений {current_time}.xlsx')
                else:
                    # сохраняем

                    finish_result.to_excel(f'{path_to_end_folder_calculate_data}/Итоговые значения {current_time}.xlsx',
                                           index=False)

                if count_errors != 0:
                    messagebox.showinfo('Эльпида Школьная отчетность',
                                        f'В некоторых файлах обнаружены ошибки!\nОбработано файлов:  {count} из {quantity_files}\n Необработанные файлы указаны в файле {path_to_end_folder_calculate_data}/Ошибки {current_time}.xlsx')
                else:
                    messagebox.showinfo('Эльпида Школьная отчетность',
                                        f'Обработка файлов успешно завершена!\nОбработано файлов:  {count} из {quantity_files}')
            else:
                messagebox.showwarning('Эльпида Школьная отчетность',
                                    f'Обработано {count} из {quantity_files} файлов.\n Причины необработки файлов указаны в файле {path_to_end_folder_calculate_data}/Ошибки {current_time}.xlsx')

    except UnboundLocalError:
        pass
    except NameError:
        messagebox.showerror('Эльпида Школьная отчетность',
                             f'Выберите шаблон,файл с данными и папку куда будут генерироваться файлы')
    except FileNotFoundError:
        messagebox.showerror('Эльпида Школьная отчетность',
                             f'Перенесите файлы, конечную папку с которой вы работете в корень диска. Проблема может быть\n '
                             f'в слишком длинном пути к обрабатываемым файлам или конечной папке.')
    except NotFile:
        messagebox.showerror('Эльпида Школьная отчетность',
                             f'В выбранной папке отсутствуют файлы Excel (xlsx))')
    except NotCorrectParams:
        messagebox.showerror('Эльпида Школьная отчетность',
                             f'Не найдено ни одного корректного адреса ячейки. Адрес ячейки должен выглядить так A5, AB14 и т.п.')
    except:
        logging.exception('AN ERROR HAS OCCURRED')
        messagebox.showerror('Эльпида Школьная отчетность',
                             'Возникла ошибка!!! Подробности ошибки в файле error.log')

if __name__ == '__main__':
    mode_text = 'No'
    name_file_params_calculate_data = 'data/Извлечение данных/Анкеты мониторинг профориентации/Параметры для подсчета анкет.xlsx'
    names_files_calculate_data = 'data/Извлечение данных/Анкеты мониторинг профориентации/Школы Мониторинг профориентации'

    # name_file_params_calculate_data = 'data/Параметры отчета Пример 1.xlsx'
    # names_files_calculate_data = 'data/Табличные отчеты Пример 1'

    # names_files_calculate_data = ''
    path_to_end_folder_calculate_data = 'data'


    extract_data_from_hard_xlsx(mode_text, name_file_params_calculate_data, names_files_calculate_data,
                                path_to_end_folder_calculate_data)