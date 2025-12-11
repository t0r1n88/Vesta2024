"""
Скрипт для подготовки списка
Очистка некорректных данных, удаление лишних пробелов
"""
from support_functions import write_df_highlighting_error_to_excel, del_sheet, write_df_to_excel_error_prep_list # функция для записи в файл Excel с автоподбором ширины
import time
import gc
import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows

import numpy as np
import xlsxwriter
from datetime import datetime
import re
from tkinter import messagebox
import logging
logging.basicConfig(
    level=logging.WARNING,
    filename="error.log",
    filemode='w',
    # чтобы файл лога перезаписывался  при каждом запуске.Чтобы избежать больших простыней. По умолчанию идет 'a'
    format="%(asctime)s - %(module)s - %(levelname)s - %(funcName)s: %(lineno)d - %(message)s",
    datefmt='%H:%M:%S',
)
import warnings
# warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
warnings.simplefilter(action='ignore', category=DeprecationWarning)
warnings.simplefilter(action='ignore', category=UserWarning)
warnings.simplefilter(action='ignore', category=FutureWarning)
pd.options.mode.chained_assignment = None

class ExceedingQuantity(Exception):
    """
    Исключение для случаев когда числа уникальных значений больше 255
    """
    pass

class NotNumberColumn(Exception):
    """
    Исключение для обработки варианта когда в таблице нет колонки с таким порядковым номером
    """
    pass

class NoMoreNumberColumn(Exception):
    """
    Исключение для обработки варианта если в строке с указанием колонок по которым нужно проверить дубликаты нет цифр

    """
    pass



def convert_to_date_prep_list(value,current_date):
    """
    Функция для конвертации строки в текст
    :param value: значение для конвертации
    :param current_date: текущая дата
    :return:
    """
    try:
        if 'Ошибка' in value:
            return value
        else:
            date_value = datetime.strptime(value, '%Y-%m-%d %H:%M:%S')
            if date_value.date() > current_date:
                string_date = datetime.strftime(date_value, '%d.%m.%Y')

                return f'Ошибка: {string_date}, превышает текущую дату. Проверьте значение или системное время на компьютере'
            return date_value
    except ValueError:
        result = re.search(r'^\d{2}\.\d{2}\.\d{4}$', value)
        if result:
            try:
                temp_date = datetime.strptime(result.group(0), '%d.%m.%Y')
                if temp_date.date() > current_date:
                    string_date = datetime.strftime(temp_date, '%d.%m.%Y')
                    return f'Ошибка: {string_date}, превышает текущую дату. Проверьте значение или системное время на компьютере'
                return temp_date
            except ValueError:
                # для случаев вида 45.09.2007
                return f'Ошибка: {value}, проверьте  правильность даты'
        else:
            # Пытаемся обработать варианты с пробелом между блоками
            value = str(value)
            result_short_yandex = re.search(r'^\d{4}-\d{2}-\d{2}$',value) # ищем сокращенный вариант яндекса
            if result_short_yandex:
                try:
                    temp_date_yandex = datetime.strptime(result_short_yandex.group(0), '%Y-%m-%d')
                    if temp_date_yandex.date() > current_date:
                        string_date = datetime.strftime(temp_date_yandex, '%d.%m.%Y')
                        return f'Ошибка: {string_date}, превышает текущую дату. Проверьте значение или системное время на компьютере'
                    return temp_date_yandex
                except ValueError:
                    # для случаев вида 45.09.2007
                    return f'Ошибка: {value}, проверьте  правильность даты'
            else:
                lst_dig = re.findall(r'\d',value)
                if len(lst_dig) != 8:
                    return f'Ошибка: {value}, проверьте  правильность даты'
                # делаем строку
                temp_date = f'{lst_dig[0]}{lst_dig[1]}.{lst_dig[2]}{lst_dig[3]}.{lst_dig[4]}{lst_dig[5]}{lst_dig[6]}{lst_dig[7]}'
                try:
                    temp_date = datetime.strptime(temp_date, '%d.%m.%Y')
                    if temp_date.date() > current_date:
                        string_date = datetime.strftime(temp_date, '%d.%m.%Y')
                        return f'Ошибка: {string_date}, превышает текущую дату. Проверьте значение или системное время на компьютере'
                    return temp_date
                except ValueError:
                    # для случаев вида 45.09.2007
                    return f'Ошибка: {value}, проверьте  правильность даты'

    except:
        return f'Ошибка: {value}, проверьте  правильность даты'





def create_doc_convert_date(cell):
    """
    Функция для конвертации даты при создании документов
    :param cell:
    :return:
    """
    try:
        if pd.isna(cell):
            return 'Ошибка: Не заполнено'
        string_date = datetime.strftime(cell, '%d.%m.%Y')
        return string_date
    except:
        return f'Ошибка - {cell}'


def capitalize_fio(value:str)->str:
    """
    Функция для применения capitalize к значениям состоящим из несколько слов разделенных пробелами
    value: значение ячейки
    """
    value = str(value)
    if value == 'Не заполнено':
        return value
    temp_lst = value.split(' ') # создаем список по пробелу
    temp_lst = list(map(str.capitalize,temp_lst))  # обрабатываем
    return ' '.join(temp_lst) #соединяем в строку


def find_english_letter(value):
    """
    Функция для поиска английских букв в ФИО
    :param value: строка ФИО
    :return:
    """
    result = re.findall(r'[a-zA-Z]',value)
    if result:
        english_let = ';'.join(result)
        return f'Ошибка: обнаружены символы английского алфавита: {english_let} в слове {value}'
    else:
        return value

def prepare_fio_text_columns(df:pd.DataFrame,lst_columns:list)->pd.DataFrame:
    """
    Функция для очистки текстовых колонок c данными ФИО
    df: датафрейм для обработки
    lst_columns: список колонок которые нужно обработать
    """
    prepared_columns_lst = [] # список для колонок содержащих слова Фамилия,Имя,Отчество, ФИО
    for fio_column in lst_columns:
        for name_column in df.columns:
            if fio_column in name_column.lower():
                prepared_columns_lst.append(name_column)
    if len(prepared_columns_lst) == 0: # проверка на случай не найденных значений
        return df

    df[prepared_columns_lst] = df[prepared_columns_lst].fillna('Ошибка: Не заполнено')
    df[prepared_columns_lst] = df[prepared_columns_lst].astype(str)
    df[prepared_columns_lst] = df[prepared_columns_lst].applymap(lambda x: x.strip() if isinstance(x, str) else x)  # применяем strip, чтобы все данные корректно вставлялись
    df[prepared_columns_lst] = df[prepared_columns_lst].applymap(lambda x:' '.join(x.split())) # убираем лишние пробелы между словами
    df[prepared_columns_lst] = df[prepared_columns_lst].applymap(capitalize_fio)  # делаем заглавными первые буквы слов а остальыне строчными

    return df

def prepare_date_column(df:pd.DataFrame,lst_columns:list)->pd.DataFrame:
    """
    Функция для обработки колонок с датами
    df: датафрейм для обработки
    lst_columns: список колонок которые нужно обработать
    """
    current_date = datetime.now().date()  # Получаем текущую дату
    prepared_columns_lst = [] # список для колонок содержащих слово дата
    for date_column in lst_columns:
        for name_column in df.columns:
            if date_column in name_column.lower():
                prepared_columns_lst.append(name_column)
    if len(prepared_columns_lst) == 0: # проверка на случай не найденных значений
        return df
    df[prepared_columns_lst] = df[prepared_columns_lst].fillna('Ошибка: Не заполнено')
    df[prepared_columns_lst] = df[prepared_columns_lst].applymap(lambda x:convert_to_date_prep_list(x,current_date)) # приводим к типу дата
    df[prepared_columns_lst] = df[prepared_columns_lst].applymap(create_doc_convert_date)  # приводим к виду ДД.ММ.ГГГГ
    return df

def prepare_snils(df:pd.DataFrame,snils:str)->pd.DataFrame:
    """
    Функция для обработки колонок со снилс
    df: датафрейм для обработки
    snils: название снилс
    """

    prepared_columns_lst = []  # список для колонок содержащих слово снилс
    for name_column in df.columns:
        if snils in name_column.lower():
            prepared_columns_lst.append(name_column)

    if len(prepared_columns_lst) == 0: # проверка на случай не найденных значений
        return df

    df[prepared_columns_lst] = df[prepared_columns_lst].applymap(check_snils)

    return df

def prepare_snils_copp(df:pd.DataFrame,snils:str)->pd.DataFrame:
    """
    Функция для обработки колонок со снилс
    df: датафрейм для обработки
    snils: название снилс
    """
    if snils not in df.columns:
        messagebox.showerror('','Не найдена колонка СНИЛС!!!')

    df['СНИЛС'] =df['СНИЛС'].apply(check_snils)
    return df



def check_snils(snils):
    """
    Функция для приведения значений снилс в вид ХХХ-ХХХ-ХХХ ХХ
    """
    if snils is np.nan:
        return 'Не заполнено'
    snils = str(snils)
    result = re.findall(r'\d', snils) # ищем цифры
    if len(result) == 11:
        first_group = ''.join(result[:3])
        second_group = ''.join(result[3:6])
        third_group = ''.join(result[6:9])
        four_group = ''.join(result[9:11])

        out_snils = f'{first_group}-{second_group}-{third_group} {four_group}'
        return out_snils
    else:
        return f'Ошибка: В СНИЛС должно быть 11 цифр - {snils} -{len(result)} цифр'

def prepare_inn_column(df:pd.DataFrame,lst_columns:list)->pd.DataFrame:
    """
    Функция для обработки колонок со снилс
    df: датафрейм для обработки
    lst_columns: список колонок с ИНН
    """

    prepared_columns_lst = [] # список для колонок содержащих слово дата
    for inn_column in lst_columns:
        for name_column in df.columns:
            if inn_column in name_column.lower():
                prepared_columns_lst.append(name_column)
    if len(prepared_columns_lst) == 0: # проверка на случай не найденных значений
        return df

    df[prepared_columns_lst] = df[prepared_columns_lst].applymap(check_inn) # обрабатываем инн
    return df


def check_inn(inn):
    """
    Функция для приведения значений снилс в вид 12 цифр
    """
    if inn is np.nan:
        return 'Ошибка: Не заполнено'
    inn = str(inn)
    result = re.findall(r'\d', inn) # ищем цифры
    if len(result) == 12:
        return ''.join(result)
    else:
        return f'Ошибка: (ИНН физлица состоит из 12 цифр)- {inn} -{len(inn)} цифр'

def prepare_passport_column(df:pd.DataFrame)->pd.DataFrame:
    """
    Функция для обработки колонок серия и номер паспорта
    df: датафрейм для обработки
    series_passport: значение для поиска колонкок с содержащей серию паспорта
    number_passport: значение для поиска колонкок с содержащей серию паспорта
    code_passport: значение для поиска колонкок с содержащей код подразделения

    """
    prepared_columns_series_lst = [] # список для колонок содержащих слова серия паспорт
    prepared_columns_number_lst = [] # список для колонок содержащих слова номер паспорт
    prepared_columns_code_lst = [] # список для колонок содержащих слова код подразд
    pattern_series = re.compile(r"(?=.*серия)(?=.*паспорт)") # паттерн для серии паспорта
    pattern_number = re.compile(r"(?=.*номер)(?=.*паспорт)") # паттерн для номера паспорта
    pattern_code = re.compile(r"(?=.*код)(?=.*подразд)") # паттерн для кода подразделения
    for name_column in df.columns:
        result_series = re.search(pattern_series,name_column.lower()) # ищем по паттерну серию
        if result_series:
            prepared_columns_series_lst.append(name_column)
        result_number = re.search(pattern_number,name_column.lower()) # ищем по паттерну номер
        if result_number:
            prepared_columns_number_lst.append(name_column)
        result_code =   re.search(pattern_code,name_column.lower()) # ищем по паттерну код подразделения
        if result_code:
            prepared_columns_code_lst.append(name_column)


    if len(prepared_columns_series_lst) != 0:
        df[prepared_columns_series_lst] = df[prepared_columns_series_lst].applymap(check_series_passport)  # обрабатываем серию паспорта

    if len(prepared_columns_number_lst) != 0:
        df[prepared_columns_number_lst] = df[prepared_columns_number_lst].applymap(check_number_passport)  # обрабатываем номер паспорта

    if len(prepared_columns_code_lst) != 0:
        df[prepared_columns_code_lst] = df[prepared_columns_code_lst].applymap(check_code_passport)  # обрабатываем код подразделения

    return df

def check_series_passport(series:str)->str:
    """
    Функция для проверки серии паспорта, должно быть 4 цифры
    """
    if series is np.nan:
        return 'Ошибка: Не заполнено'
    series = str(series)
    result = re.findall(r'\d', series) # ищем цифры
    if len(result) == 4:
        return ''.join(result)
    else:
        return f'Ошибка: в серии паспорта должно быть 4 цифры - {series}'

def check_number_passport(number:str)->str:
    """
    Функция для проверки номера паспорта, должно быть 6 цифр
    """
    if number is np.nan:
        return 'Ошибка: Не заполнено'
    number = str(number)
    result = re.findall(r'\d', number) # ищем цифры
    if len(result) == 6:
        return ''.join(result)
    else:
        return f'Ошибка: в номере паспорта должно быть 6 цифр - {number}'

def check_code_passport(code:str)->str:
    """
    Функция для проверки номера паспорта, должно быть 6 цифр
    """
    if code is np.nan:
        return 'Не заполнено'
    code = str(code)
    result = re.findall(r'\d', code) # ищем цифры
    if len(result) == 6:
        first_group = ''.join(result[:3])
        second_group = ''.join(result[3:6])
        return f'{first_group}-{second_group}'
    else:
        return f'Ошибка: в коде подразделения должно быть 6 цифр в формате XXX-XXX - {code}'

def prepare_phone_columns(df:pd.DataFrame,phone_text:str) ->pd.DataFrame:
    """
    Функция для очистки номеров телефонов от пробельных символов и букв
    """
    # pattern = r'[a-zA-Zа-яА-Я\s.]'
    pattern = r'\D' # удаляем  все кроме цифр
    prepared_phone_columns = [] # лист для колонок с телефонами
    # собираем названия колонок содержащих слово телефон
    for name_column in df.columns:
        if phone_text in name_column.lower():
            prepared_phone_columns.append(name_column)

    if len(prepared_phone_columns) == 0:
        return df

    df[prepared_phone_columns] = df[prepared_phone_columns].applymap(lambda x:check_phone_number(x,pattern))
    return df

def check_phone_number(phone:str,pattern:str)->str:
    """
    Функция для очистки значения номера телефона от пробельных символов,букв и точки
    """
    if phone is np.nan:
        return 'Ошибка: Не заполнено'
    phone = str(phone)
    clean_phone = re.sub(pattern,'',phone)
    return clean_phone


def prepare_email_columns(df:pd.DataFrame,second_option:str)->pd.DataFrame:
    """
    Функция для обработки колонок серия и номер паспорта
    df: датафрейм для обработки
    second_option: значение для поиска колонкок с содержащей слово e-mail
    """
    prepared_columns_email_set = set() # множество для колонок содержащих слова электрон почта e-mail
    # это нужно для обработки случаем когда колонка называется Электронная почта(email)
    pattern_first_option = re.compile(r"(?=.*электрон)(?=.*почта)") # паттерн для слов электрон почта
    for name_column in df.columns:
        result_first_option = re.search(pattern_first_option,name_column.lower()) # ищем по паттерну электрон почта
        if result_first_option:
            prepared_columns_email_set.add(name_column)
        if second_option in name_column:
            prepared_columns_email_set.add(name_column)
    prepared_columns_email_lst = list(prepared_columns_email_set) # превращаем в список

    if len(prepared_columns_email_lst) == 0:
        return df
    df[prepared_columns_email_lst] = df[prepared_columns_email_lst].fillna('Ошибка: Не заполнено')
    df[prepared_columns_email_lst] = df[prepared_columns_email_lst].applymap(lambda x:re.sub(r'\s','',x) if x !='Ошибка: Не заполнено' else x)

    return df


def check_mixing(value:str):
    """
    Функция для проверки слова на смешение алфавитов
    """
    # ищем буквы русского и английского алфавита
    russian_letters = re.findall(r'[а-яА-ЯёЁ]',value)
    english_letters = re.findall(r'[a-zA-Z]',value)
    # если найдены и те и те
    if russian_letters and english_letters:
        # если русских букв больше то указываем что в русском слове встречаются английские буквы
        if len(russian_letters) > len(english_letters):
            return (f'Ошибка: в слове {value} найдены английские буквы: {",".join(english_letters)}')
        elif len(russian_letters) < len(english_letters):
            # если английских букв больше то указываем что в английском слове встречаются русские буквы
            return (f'Ошибка: в слове {value} найдены русские буквы: {",".join(russian_letters)}')
        else:
            # если букв поровну то просто выводим их список
            return (f'Ошибка: в слове {value} найдены русские буквы: {",".join(russian_letters)} и английские буквы: {";".join(english_letters)}')
    else:
        # если слово состоит из букв одного алфавита
        return False


def find_mixing_alphabets(cell):
    """
    Функция для нахождения случаев смешения когда английские буквы используются в русском слове и наоборот
    """
    if isinstance(cell,str):
        lst_word = re.split(r'\W',cell) # делим по не буквенным символам
        lst_result = list(map(check_mixing,lst_word)) # ищем смешения
        lst_result = [value for value in lst_result if value] # отбираем найденые смешения если они есть
        if lst_result:
            return f'Ошибка: в тексте {cell} найдено смешение русского и английского: {"; ".join(lst_result)}'
        else:
            return cell
    else:
        return cell

def prepare_entry_str(raw_str:str,pattern:str,repl_str:str,sep_lst:str)->list:
    """
    Функция для очистки строки от лишних символов и уменьшения на единицу (для нумерации с нуля)
    :param raw_str: обрабатываемая строка
    :param pattern: паттерн для замены символов
    :param repl_str: строка на которую нужно заменять символы
    :param sep_lst: разделитель по которому будет делиться список
    :return: список
    """
    raw_str = str(raw_str).replace('.',',')
    number_column_folder_structure = re.sub(pattern,repl_str,raw_str) # убираем из строки все лишние символы
    lst_number_column_folder_structure = number_column_folder_structure.split(sep_lst) # создаем список по запятой
    # отбрасываем возможные лишние элементы из за лишних запятых
    lst_number_column_folder_structure = [value for value in lst_number_column_folder_structure if value]
    # заменяем 0 на единицу
    lst_number_column_folder_structure = ['1' if x == '0' else x for x in lst_number_column_folder_structure]
    # Превращаем в числа и отнимаем 1 чтобы соответствовать индексам питона
    lst_number_column_folder_structure = list(map(lambda x:int(x)-1,lst_number_column_folder_structure))
    return lst_number_column_folder_structure


def convert_columns_to_str(df, number_columns):
    """
    Функция для конвертации указанных столбцов в строковый тип и очистки от пробельных символов в начале и конце
    """

    for column in number_columns:  # Перебираем список нужных колонок
        try:
            df.iloc[:, column] = df.iloc[:, column].astype(str)
            # Очищаем колонку от пробельных символов с начала и конца
            df.iloc[:, column] = df.iloc[:, column].apply(lambda x: x.strip())
        except IndexError:
            messagebox.showerror('Веста Обработка таблиц и создание документов',
                                 'Проверьте порядковые номера колонок которые вы хотите обработать.')




def prepare_list(file_data:str,path_end_folder:str,checkbox_dupl:str,checkbox_mix_alphabets:str,checkbox_many_dupl:str,number_dupl_columns):
    """
    file_data : путь к файлу который нужно преобразовать
    path_end_folder :  путь к конечной папке
    checkbox_dupl: Проверять на дубликаты или нет. Yes or No
    checkbox_mix_alphabets: Проверять на смешение русских и английских букв или нет. Yes or No
    checkbox_many_dupl: Проверять на дубликаты несколько колонок или нет. Yes or No
    number_dupl_columns: Порядковые номера колонок по которым нужно проверить дубликаты
    """
    try:
        try:
            df = pd.read_excel(file_data,dtype=str) # считываем датафрейм
        except:
            messagebox.showerror('Веста Обработка таблиц и создание документов',
                                 f'Не удалось обработать файл. Возможно файл поврежден')
        df.columns = list(map(str,list(df.columns))) # делаем названия колонок строкововыми
        # проверяем корректность введенных номеров колонок
        if checkbox_many_dupl == 'Yes':
            # очищаем строку от лишних символов и превращаем в список номеров колонок
            lst_number_dupl_cols = prepare_entry_str(number_dupl_columns, r'[^\d,]', '', ',')
            if len(lst_number_dupl_cols) == 0:
                raise NoMoreNumberColumn
            else:
                # проверяем чтобы номер колонки не превышал количество колонок в датафрейме
                for number_column in lst_number_dupl_cols:
                    if number_column > df.shape[1]-1:
                        raise NotNumberColumn


        # очищаем все строковые значения от пробелов в начале и конце
        df = df.applymap(lambda x: x.strip() if isinstance(x, str) else x)
        # заменяем пробельные символы на пробел, чтобы убрать лишние пробелы
        df = df.applymap(lambda x: re.sub(r'\s+', ' ', x) if isinstance(x, str) else x)
        # обрабатываем колонки с фио
        part_fio_columns = ['фамилия','имя','отчество','фио'] # колонки с типичными названиями
        df = prepare_fio_text_columns(df,part_fio_columns) # очищаем колонки с фио

        # обрабатываем колонки содержащими слово дата
        part_date_columns = ['дата']
        df = prepare_date_column(df,part_date_columns)

        # обрабатываем колонки со снилс
        snils = 'снилс'
        df = prepare_snils(df, snils)

        # обрабатываем колонки с ИНН
        part_inn_columns = ['инн']
        df = prepare_inn_column(df,part_inn_columns)

        # обрабатываем колонки данные паспорта
        df = prepare_passport_column(df)

        # обрабатываем  колонки с номера телефонов
        phone = 'телефон'
        df = prepare_phone_columns(df, phone)

        # очищаем email от пробельных символов
        second_option = 'e-mail' # слова электрон и почта используются внутри функции
        df = prepare_email_columns(df,second_option)

        # Ищем смешение английских и русских букв
        df = df.applymap(find_mixing_alphabets)  # ищем смешения

        # получаем время
        t = time.localtime()
        current_time = time.strftime('%H_%M_%S', t)

        if checkbox_dupl == 'Yes':
            """
              Создаем список дубликатов
              """
            dct_dupl_df = dict() # создаем словарь для хранения названия и датафрейма
            lst_name_columns = list(df.columns)  # получаем список колонок
            used_name_sheet = []  # список для хранения значений которые уже были использованы
            if len(lst_name_columns) >= 253:  # проверяем количество колонок которые могут созданы
                raise ExceedingQuantity
            #
            wb = xlsxwriter.Workbook(f'{path_end_folder}/Дубликаты в каждой колонке {current_time}.xlsx',{'constant_memory': True,'nan_inf_to_errors': True})  # создаем файл
            for idx, value in enumerate(lst_name_columns):
                temp_df = df[df[value].duplicated(keep=False)]  # получаем дубликаты
                if temp_df.shape[0] == 0:
                    continue

                short_value = value[:20]  # получаем обрезанное значение
                short_value = re.sub(r'[\r\b\n\t\[\]\'+()<> :"?*|\\/]', '_', short_value)

                if short_value in used_name_sheet:
                    short_value = f'{short_value}_{idx}'  # добавляем окончание

                temp_df = temp_df.sort_values(by=value)
                #     # Добавляем +2 к индексу чтобы отобразить точную строку
                temp_df.insert(0, '№ строки дубликата ', list(map(lambda x: x + 2, list(temp_df.index))))
                temp_df.replace(np.nan, None,inplace=True) # для того чтобы в пустых ячейках ничего не отображалось
                dct_dupl_df[short_value] = temp_df

            for name_sheet, dupl_df in dct_dupl_df.items():
                data_lst = dupl_df.values.tolist() # преобразуем в список
                wb_name_sheet = wb.add_worksheet(name_sheet) # создаем лист
                used_name_sheet.append(name_sheet) # добавляем в список использованных названий
                # Запись заголовков
                headers = list(dupl_df.columns)
                for col, header in enumerate(headers):
                    wb_name_sheet.write(0, col, header)

                # Запись данных
                for row, data_row in enumerate(data_lst):
                    for col, cell_value in enumerate(data_row):
                        wb_name_sheet.write(row + 1, col, cell_value)

            # закрываем
            wb.close()

        if checkbox_mix_alphabets == 'Yes':
            dct_mix_df = dict()
            check_word = 'найдено смешение русского и английского:' # фраза по которой будет производится отбор
            lst_name_columns = list(df.columns)  # получаем список колонок
            used_name_sheet = []  # список для хранения значений которые уже были использованы
            if len(lst_name_columns) >= 253:  # проверяем количество колонок которые могут созданы
                raise ExceedingQuantity
            #
            wb_mix = xlsxwriter.Workbook(f'{path_end_folder}/Смешения русских и английских букв в словах {current_time}.xlsx',{'constant_memory': True,'nan_inf_to_errors': True})  # создаем файл

            for idx, value in enumerate(lst_name_columns):
                temp_df = df[df[value].astype(str).str.contains(check_word)]  # получаем строки где есть сочетание
                if temp_df.shape[0] == 0:
                    continue

                short_value = value[:20]  # получаем обрезанное значение
                short_value = re.sub(r'[\r\b\n\t\[\]\'+()<> :"?*|\\/]', '_', short_value)

                if short_value in used_name_sheet:
                    short_value = f'{short_value}_{idx}'  # добавляем окончание

                temp_df = temp_df.sort_values(by=value)
                #     # Добавляем +2 к индексу чтобы отобразить точную строку
                temp_df.insert(0, '№ строки смешения ', list(map(lambda x: x + 2, list(temp_df.index))))
                temp_df.replace(np.nan, None,inplace=True) # для того чтобы в пустых ячейках ничего не отображалось
                dct_mix_df[short_value] = temp_df

            for name_sheet, mix_df in dct_mix_df.items():
                data_lst = mix_df.values.tolist() # преобразуем в список
                wb_name_sheet = wb_mix.add_worksheet(name_sheet) # создаем лист
                used_name_sheet.append(name_sheet) # добавляем в список использованных названий
                # Запись заголовков
                headers = list(mix_df.columns)
                for col, header in enumerate(headers):
                    wb_name_sheet.write(0, col, header)

                # Запись данных
                for row, data_row in enumerate(data_lst):
                    for col, cell_value in enumerate(data_row):
                        wb_name_sheet.write(row + 1, col, cell_value)

            wb_mix.close()

        # Создаем файл с количеством по каждой колонке
        wb_stat = openpyxl.Workbook()
        main_df = df.copy() # делаем копию
        main_df['Для подсчета'] = 1
        # Создаем листы
        for idx, name_column in enumerate(main_df.columns,1):
            # Делаем короткое название не более 30 символов
            if name_column == 'Для подсчета':
                continue
            wb_stat.create_sheet(title=str(idx), index=idx)

        for idx, name_column in enumerate(main_df.columns,1):
            group_df = main_df.groupby([name_column]).agg({'Для подсчета': 'sum'})
            group_df.columns = ['Количество']

            # Сортируем по убыванию
            group_df.sort_values(by=['Количество'], inplace=True, ascending=False)
            group_df.loc['Итого'] = group_df['Количество'].sum()
            if name_column == 'Для подсчета':
                continue

            for r in dataframe_to_rows(group_df, index=True, header=True):
                if len(r) != 1:
                    wb_stat[str(idx)].append(r)
            wb_stat[str(idx)].column_dimensions['A'].width = 50

        # Удаляем листы
        del_sheet(wb_stat, ['Sheet', 'Для подсчета'])
        wb_stat.save(f'{path_end_folder}/Количество {current_time}.xlsx')

        # Если поставлен чекбокс, то проверяем несколько колонок на дубликаты
        if checkbox_many_dupl == 'Yes':
            lst_dupl_name_columns = [df.columns[idx_column] for idx_column in lst_number_dupl_cols ]
            dct_many_dupl_df = dict() # словарь для хранения дубликатов по нескольким колонкам
            for i in range(len(lst_number_dupl_cols)):
                lst_cols_for_many_dupl_df = ['№ строки дубликата','ID_дубликата']
                lst_cols_for_many_dupl_df.extend(list(df.columns))
                base_main_dupl_df = df.copy()  # создаем базовый датафрейм для дубликатов по многим колонкам
                main_dupl_df = pd.DataFrame(columns=lst_cols_for_many_dupl_df)
                if i == 0:
                    # temp_many_df = base_main_dupl_df.iloc[:,lst_number_dupl_cols]
                    dupl_many_df = base_main_dupl_df[base_main_dupl_df[lst_dupl_name_columns].duplicated(keep=False)]
                    dupl_many_df.insert(0, '№ строки дубликата', list(map(lambda x: x + 2, list(dupl_many_df.index))))
                    if len(dupl_many_df) != 0:
                        # Конвертируем нужные нам колонки в str
                        convert_columns_to_str(dupl_many_df, lst_number_dupl_cols)
                        # создаем датафреймы из колонок выбранных для объединения, такой способо связан с тем, что
                        # при использовании sum числа в строковом виде превращаются в числа
                        # Создаем в каждом датафрейме колонку с айди путем склеивания всех нужных колонок в одну строку
                        dupl_many_df.insert(1,'ID_дубликата',dupl_many_df[lst_dupl_name_columns].apply(lambda x: '_'.join(x), axis=1))
                        # dupl_many_df['ID_дубликата'] = dupl_many_df[lst_dupl_name_columns].apply(lambda x: ''.join(x), axis=1)
                        lst_id_dupl = dupl_many_df['ID_дубликата'].unique() # уникальные дубликаты
                        for id_dupl in lst_id_dupl:
                            temp_dupl_df = dupl_many_df[dupl_many_df['ID_дубликата'] == id_dupl]
                            temp_dupl_df = temp_dupl_df.sort_values(by='ID_дубликата')
                            temp_dupl_df.loc['Граница'] = ''
                            main_dupl_df = pd.concat([main_dupl_df, temp_dupl_df])

                        dct_many_dupl_df[f'{len(lst_number_dupl_cols)}'] = main_dupl_df


                else:
                    dupl_many_df = base_main_dupl_df[base_main_dupl_df[lst_dupl_name_columns[:-i]].duplicated(keep=False)]
                    dupl_many_df.insert(0, '№ строки дубликата', list(map(lambda x: x + 2, list(dupl_many_df.index))))
                    if len(dupl_many_df) != 0:
                        # Конвертируем нужные нам колонки в str
                        convert_columns_to_str(dupl_many_df, lst_number_dupl_cols[:-i])
                        # создаем датафреймы из колонок выбранных для объединения, такой способо связан с тем, что
                        # при использовании sum числа в строковом виде превращаются в числа
                        # Создаем в каждом датафрейме колонку с айди путем склеивания всех нужных колонок в одну строку
                        dupl_many_df.insert(1, 'ID_дубликата',dupl_many_df[lst_dupl_name_columns[:-i]].apply(lambda x: '_'.join(x), axis=1))

                        lst_id_dupl = dupl_many_df['ID_дубликата'].unique() # уникальные дубликаты

                        for id_dupl in lst_id_dupl:
                            temp_dupl_df = dupl_many_df[dupl_many_df['ID_дубликата'] == id_dupl]
                            temp_dupl_df = temp_dupl_df.sort_values(by='ID_дубликата')
                            temp_dupl_df.loc['Граница'] = ''
                            main_dupl_df = pd.concat([main_dupl_df, temp_dupl_df])

                        dct_many_dupl_df[f'{len(lst_number_dupl_cols[:-i])}'] = main_dupl_df


            file_error_wb = write_df_to_excel_error_prep_list(dct_many_dupl_df, write_index=False)
            file_error_wb = del_sheet(file_error_wb, ['Sheet', 'Sheet1', 'Для подсчета'])
            file_error_wb.save(f'{path_end_folder}/Дубли по нескольким колонкам {current_time}.xlsx')





        # сохраняем основной файл

        dct_df = {'Лист1': df}
        write_index = False
        wb_main = write_df_highlighting_error_to_excel(dct_df, write_index)
        wb_main = del_sheet(wb_main,['Sheet', 'Sheet1', 'Для подсчета'])

        if file_data.endswith('.xlsx'):
            name_file = file_data.split('.xlsx')[0]  # получаем путь без расширения
        else:
            name_file = file_data.split('.xlsm')[0]  # получаем путь без расширения
        name_file = name_file.split('/')[-1]
        wb_main.save(f'{path_end_folder}/Обработанный {name_file} {current_time}.xlsx')

        # Сохраняем датафрейм с ошибками разделенными по листам в соответсвии с колонками
        dct_sheet_error_df = dict()  # создаем словарь для хранения названия и датафрейма
        used_name_sheet = set()  # множество для хранения значений которые уже были использованы

        lst_name_columns = [name_cols for name_cols in df.columns if
                            'Unnamed' not in name_cols]  # получаем список колонок

        for idx, value in enumerate(lst_name_columns):
            # получаем ошибки
            temp_df = df[df[value].astype(str).str.contains('Ошибка')]  # фильтруем
            if temp_df.shape[0] == 0:
                continue

            temp_df = temp_df[value].to_frame()  # оставляем только одну колонку

            temp_df.insert(0, '№ строки с ошибкой в исходном файле', list(map(lambda x: x + 2, list(temp_df.index))))
            short_value = value[:27]  # получаем обрезанное значение
            short_value = re.sub(r'[\[\]\'+()<> :"?*|\\/]', '_', short_value)

            if short_value.lower() in used_name_sheet:
                short_value = f'{short_value}_{idx}'  # добавляем окончание
            used_name_sheet.add(short_value.lower())

            dct_sheet_error_df[short_value] = temp_df
        if len(dct_sheet_error_df) != 0:
            file_error_wb = write_df_to_excel_error_prep_list(dct_sheet_error_df, write_index=False)
            file_error_wb = del_sheet(file_error_wb, ['Sheet', 'Sheet1', 'Для подсчета'])
            file_error_wb.save(f'{path_end_folder}/Ошибки {name_file}.xlsx')
        else:
            file_error_wb = openpyxl.Workbook()
            file_error_wb.save(f'{path_end_folder}/Ошибок {name_file}.xlsx')

    except UnboundLocalError:
        pass

    except NoMoreNumberColumn:
        messagebox.showerror('Веста Обработка таблиц и создание документов',
                             f'Введите числа через запятую. Например 3,8,25,1')
        logging.exception('AN ERROR HAS OCCURRED')

    except NotNumberColumn:
        messagebox.showerror('Веста Обработка таблиц и создание документов',
                             f'Введите числа не превышающие количество колонок в таблице которую вы хотите обработать')
        logging.exception('AN ERROR HAS OCCURRED')

    except NameError:
        messagebox.showerror('Веста Обработка таблиц и создание документов',
                             f'Выберите файлы с данными и папку куда будет генерироваться файл')
        logging.exception('AN ERROR HAS OCCURRED')

    except ValueError as e:
        messagebox.showerror('Веста Обработка таблиц и создание документов',
                             f'Ошибка при обработке значения {e.args}')
        logging.exception('AN ERROR HAS OCCURRED')

    except FileNotFoundError:
        messagebox.showerror('Веста Обработка таблиц и создание документов',
                             f'Перенесите файлы, конечную папку с которой вы работете в корень диска. Проблема может быть\n '
                             f'в слишком длинном пути к обрабатываемым файлам или конечной папке.')
    else:
        messagebox.showinfo('Веста Обработка таблиц и создание документов', 'Данные успешно обработаны')


if __name__ == '__main__':
    # file_data_main = 'data/Обработка списка/Список студентов военкомат.xlsx'
    file_data_main = 'data/Обработка списка/Список студентов военкомат.xlsx'
    file_data_main = 'data/Обработка списка/выверка.xlsx'
    path_end_main = 'data/result'
    checkbox_main_dupl = 'Yes'
    checkbox_main_mix_alphabets = 'Yes'
    checkbox_main_many_dupl_cols = 'Yes'
    main_lst_dupl_columns = '17,9,11,12,14'
    start_time = time.time()
    prepare_list(file_data_main,path_end_main,checkbox_main_dupl,checkbox_main_mix_alphabets,checkbox_main_many_dupl_cols,main_lst_dupl_columns)
    end_time = time.time()
    execution_time = end_time - start_time
    print(f"Время выполнения: {execution_time} секунд")
    print('Lindy Booth')

