"""
Вспомогательные функции
"""
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill
import re
from copy import copy
import datetime


class ExceedingQuantity(Exception):
    """
    Исключение для случаев когда числа уникальных значений больше 255
    """
    pass




def write_df_to_excel(dct_df:dict,write_index:bool)->openpyxl.Workbook:
    """
    Функция для записи датафрейма в файл Excel
    :param dct_df: словарь где ключе это название создаваемого листа а значение датафрейм который нужно записать
    :param write_index: нужно ли записывать индекс датафрейма True or False
    :return: объект Workbook с записанными датафреймами
    """
    wb = openpyxl.Workbook() # создаем файл
    count_index = 0 # счетчик индексов создаваемых листов
    for name_sheet,df in dct_df.items():
        wb.create_sheet(title=name_sheet,index=count_index) # создаем лист
        # записываем данные в лист
        for row in dataframe_to_rows(df,index=write_index,header=True):
            wb[name_sheet].append(row)
        # ширина по содержимому
        # сохраняем по ширине колонок
        for column in wb[name_sheet].columns:
            max_length = 0
            column_name = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            wb[name_sheet].column_dimensions[column_name].width = adjusted_width
        count_index += 1
    # удаляем лишний лист
    if len(wb.sheetnames) >= 2 and 'Sheet' in wb.sheetnames:
        del wb['Sheet']
    return wb

def convert_to_int(value):
    """
    Функция для конвертации значения в инт
    :param value:
    :return:
    """
    try:
        return int(value)
    except:
        return 0

def convert_to_float(value):
    try:
        return float(value)
    except:
        return 0


def convert_to_date(value):
    """
    Функция для конвертации строки в текст
    :param value: значение для конвертации
    :return:
    """
    try:
        date_value = datetime.datetime.strptime(value, '%Y-%m-%d %H:%M:%S')
        return date_value
    except ValueError:
        result = re.search(r'^\d{2}\.\d{2}\.\d{4}$', value)
        if result:
            try:
                temp_date = datetime.datetime.strptime(result.group(0), '%d.%m.%Y')
                return temp_date
            except ValueError:
                # для случаев вида 45.09.2007
                return f''
        else:
            # Пытаемся обработать варианты с пробелом или лишними точками между блоками
            value = str(value)
            lst_dig = re.findall(r'\d',value)
            if len(lst_dig) != 8:
                return f''
            # делаем строку
            temp_date = f'{lst_dig[0]}{lst_dig[1]}.{lst_dig[2]}{lst_dig[3]}.{lst_dig[4]}{lst_dig[5]}{lst_dig[6]}{lst_dig[7]}'
            try:
                temp_date = datetime.datetime.strptime(temp_date, '%d.%m.%Y')
                return temp_date
            except ValueError:
                # для случаев вида 45.09.2007
                return f''

    except:
        return f''






def create_doc_convert_date(cell):
    """
    Функция для конвертации даты при создании документов
    :param cell:
    :return:
    """
    try:
        string_date = datetime.datetime.strftime(cell, '%d.%m.%Y')
        return string_date
    except:
        return ''


def write_df_big_dct_to_excel(dct_df: dict, write_index: bool) -> openpyxl.Workbook:
    """
    Функция для записи датафрейма в файл Excel
    :param dct_df: словарь где ключе это название создаваемого листа а значение датафрейм который нужно записать
    :param write_index: нужно ли записывать индекс датафрейма True or False
    :return: объект Workbook с записанными датафреймами
    """
    wb = openpyxl.Workbook()  # создаем файл
    count_index = 0  # счетчик индексов создаваемых листов
    used_name_sheet = set()  # множество для хранения значений которые уже были использованы
    if len(dct_df) >= 253:
        raise ExceedingQuantity
    for name_sheet, df in dct_df.items():
        short_name_sheet = name_sheet[:20]  # получаем обрезанное значение
        short_name_sheet = re.sub(r'[\[\]\'+()<> :"?*|\\/]', '_', short_name_sheet)
        if short_name_sheet.lower() in used_name_sheet:
            short_name_sheet = f'{short_name_sheet}_{count_index}'  # добавляем окончание

        wb.create_sheet(title=short_name_sheet, index=count_index)  # создаем лист
        used_name_sheet.add(short_name_sheet.lower()) # добавляем в список использованных названий
        # записываем данные в лист
        none_check = None  # чекбокс для проверки наличия пустой первой строки, такое почему то иногда бывает
        for row in dataframe_to_rows(df, index=write_index, header=True):
            if len(row) == 1 and not row[0]:  # убираем пустую строку
                none_check = True
                wb[short_name_sheet].append(row)
            else:
                wb[short_name_sheet].append(row)
        if none_check:
            wb[short_name_sheet].delete_rows(2)

        # ширина по содержимому
        # сохраняем по ширине колонок
        for column in wb[short_name_sheet].columns:
            max_length = 0
            column_name = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            wb[short_name_sheet].column_dimensions[column_name].width = adjusted_width
        count_index += 1

        column_number = 0  # номер колонки
        # Создаем стиль шрифта и заливки
        font = Font(color='FF000000')  # Черный цвет
        fill = PatternFill(fill_type='solid', fgColor='ffa500')  # Оранжевый цвет
        for row in wb[short_name_sheet].iter_rows(min_row=1, max_row=wb[short_name_sheet].max_row,
                                            min_col=column_number, max_col=df.shape[1] + 1):  # Перебираем строки
            if 'Итого' in str(row[column_number].value):  # делаем ячейку строковой и проверяем наличие слова Статус_
                for cell in row:  # применяем стиль если условие сработало
                    cell.font = font
                    cell.fill = fill

    return wb



def del_sheet(wb: openpyxl.Workbook, lst_name_sheet: list) -> openpyxl.Workbook:
    """
    Функция для удаления лишних листов из файла
    :param wb: объект таблицы
    :param lst_name_sheet: список удаляемых листов
    :return: объект таблицы без удаленных листов
    """
    for del_sheet in lst_name_sheet:
        if del_sheet in wb.sheetnames:
            del wb[del_sheet]

    return wb



"""
Функции для копирования файлов вместе с форматированием и формулами
"""
def copy_sheet(source_sheet, target_sheet):
    copy_cells(source_sheet, target_sheet)  # copy all the cell values and styles
    copy_sheet_attributes(source_sheet, target_sheet)

def copy_sheet_attributes(source_sheet, target_sheet):
    # функция копирования взята отсюда https://stackoverflow.com/questions/42344041/how-to-copy-worksheet-from-one-workbook-to-another-one-using-openpyxl
    target_sheet.sheet_format = copy(source_sheet.sheet_format)
    target_sheet.sheet_properties = copy(source_sheet.sheet_properties)
    target_sheet.merged_cells = copy(source_sheet.merged_cells)
    target_sheet.page_margins = copy(source_sheet.page_margins)
    target_sheet.freeze_panes = copy(source_sheet.freeze_panes)

    # set row dimensions
    for rn in range(len(source_sheet.row_dimensions)):
        target_sheet.row_dimensions[rn] = copy(source_sheet.row_dimensions[rn])

    if source_sheet.sheet_format.defaultColWidth is None:
        pass
    else:
        target_sheet.sheet_format.defaultColWidth = copy(source_sheet.sheet_format.defaultColWidth)

    # set specific column width and hidden property
    for key, value in source_sheet.column_dimensions.items():
        target_sheet.column_dimensions[key].min = copy(source_sheet.column_dimensions[key].min)
        target_sheet.column_dimensions[key].max = copy(source_sheet.column_dimensions[key].max)
        target_sheet.column_dimensions[key].width = copy(source_sheet.column_dimensions[key].width)
        target_sheet.column_dimensions[key].hidden = copy(source_sheet.column_dimensions[key].hidden)

def copy_cells(source_sheet, target_sheet):
    for (row, col), source_cell in source_sheet._cells.items():
        target_cell = target_sheet.cell(column=col, row=row)

        # Копируем значение или формулу
        if source_cell.data_type == 'f':  # Если это формула
            target_cell.value = source_cell.value  # Копируем формулу
        else:
            target_cell._value = source_cell._value

        target_cell.data_type = source_cell.data_type

        if source_cell.has_style:
            target_cell.font = copy(source_cell.font)
            target_cell.border = copy(source_cell.border)
            target_cell.fill = copy(source_cell.fill)
            target_cell.number_format = copy(source_cell.number_format)
            target_cell.protection = copy(source_cell.protection)
            target_cell.alignment = copy(source_cell.alignment)

        if source_cell.hyperlink:
            target_cell._hyperlink = copy(source_cell.hyperlink)

        if source_cell.comment:
            target_cell.comment = copy(source_cell.comment)



def write_df_highlighting_error_to_excel(dct_df: dict, write_index: bool) -> openpyxl.Workbook:
    """
    Функция для записи датафрейма c данными и заливкой ошибок цветом
    :param dct_df: словарь где ключе это название создаваемого листа а значение датафрейм который нужно записать
    :param write_index: нужно ли записывать индекс датафрейма True or False
    :return: объект Workbook с записанными датафреймами
    """
    wb = openpyxl.Workbook()  # создаем файл
    count_index = 0  # счетчик индексов создаваемых листов
    used_name_sheet = set()  # множество для хранения значений которые уже были использованы
    if len(dct_df) >= 253:
        raise ExceedingQuantity
    for name_sheet, df in dct_df.items():
        short_name_sheet = name_sheet[:20]  # получаем обрезанное значение
        short_name_sheet = re.sub(r'[\[\]\'+()<> :"?*|\\/]', '_', short_name_sheet)
        if short_name_sheet.lower() in used_name_sheet:
            short_name_sheet = f'{short_name_sheet}_{count_index}'  # добавляем окончание

        wb.create_sheet(title=short_name_sheet, index=count_index)  # создаем лист
        used_name_sheet.add(short_name_sheet.lower()) # добавляем в список использованных названий
        # записываем данные в лист
        none_check = None  # чекбокс для проверки наличия пустой первой строки, такое почему то иногда бывает
        for row in dataframe_to_rows(df, index=write_index, header=True):
            if len(row) == 1 and not row[0]:  # убираем пустую строку
                none_check = True
                wb[short_name_sheet].append(row)
            else:
                wb[short_name_sheet].append(row)
        if none_check:
            wb[short_name_sheet].delete_rows(2)

        # ширина по содержимому
        # сохраняем по ширине колонок
        for column in wb[short_name_sheet].columns:
            max_length = 0
            column_name = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            if adjusted_width > 40:
                adjusted_width = 40
            wb[short_name_sheet].column_dimensions[column_name].width = adjusted_width



        count_index += 1

        start_column_number = 0  # номер колонки
        # Создаем стиль шрифта и заливки
        font = Font(color='FF000000')  # Черный цвет
        fill = PatternFill(fill_type='solid', fgColor='ffa500')  # Оранжевый цвет
        for row in wb[short_name_sheet].iter_rows(min_row=1, max_row=wb[short_name_sheet].max_row,
                                            min_col=start_column_number, max_col=df.shape[1]):  # Перебираем строки
            for i in range(1,df.shape[1]):
                if 'Ошибка' in str(row[i].value):  # делаем ячейку строковой и проверяем наличие слова Статус_
                    for cell in row:  # применяем стиль если условие сработало
                        cell.font = font
                        cell.fill = fill

    return wb


def write_df_to_excel_error_prep_list(dct_df: dict, write_index: bool) -> openpyxl.Workbook:
    """
    Функция для записи датафрейма в файл Excel
    :param dct_df: словарь где ключе это название создаваемого листа а значение датафрейм который нужно записать
    :param write_index: нужно ли записывать индекс датафрейма True or False
    :return: объект Workbook с записанными датафреймами
    """
    if len(dct_df) >= 253:
        raise ExceedingQuantity # проверяем количество значений
    wb = openpyxl.Workbook()  # создаем файл
    count_index = 0  # счетчик индексов создаваемых листов
    for name_sheet, df in dct_df.items():
        wb.create_sheet(title=name_sheet, index=count_index)  # создаем лист
        # записываем данные в лист
        none_check = None  # чекбокс для проверки наличия пустой первой строки, такое почему то иногда бывает
        for row in dataframe_to_rows(df, index=write_index, header=True):
            if len(row) == 1 and not row[0]:  # убираем пустую строку
                none_check = True
                wb[name_sheet].append(row)
            else:
                wb[name_sheet].append(row)
        if none_check:
            wb[name_sheet].delete_rows(2)

        # ширина по содержимому
        # сохраняем по ширине колонок
        for column in wb[name_sheet].columns:
            max_length = 0
            column_name = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            # для слишком длинных результатов
            if adjusted_width > 60:
                adjusted_width = 60
            wb[name_sheet].column_dimensions[column_name].width = adjusted_width
        count_index += 1

    return wb