"""
Скрипт для сбора файлов xlsx с сохранением форматирования в один файл по листам
"""
from support_functions import copy_sheet
import openpyxl
import pandas as pd
import os
import time
import re
from tkinter import messagebox




def union_files_xlsx(data_folder:str,end_folder:str):
    """
    Функция для сбора файлов с сохранением форматирования по листам в один файл
    :param data_folder: папка с таблицами
    :param end_folder: конечная папка
    """
    try:
        count_errors = 0 # количество ошибок
        count_sheets = 1 # количество листов в сводном файле

        error_df = pd.DataFrame(
            columns=['Название файла', 'Описание ошибки'])  # датафрейм для ошибок

        t = time.localtime()
        current_time = time.strftime('%H_%M_%S', t)

        main_wb = openpyxl.Workbook() # сводный файл
        alone_sheet = main_wb[main_wb.sheetnames[0]]
        alone_sheet.title = 'alone_sheet' # переименовываем лист чтобы не возникло конфликтов

        set_used_named = set() # множество для хранения использованных коротких названий

        for dirpath, dirnames, filenames in os.walk(data_folder):
            for file in filenames:
                if not file.startswith('~$') and file.endswith('.xlsx'):
                    name_file = file.split('.xlsx')[0].strip() # имя файла
                    try:
                        print(name_file)  # обрабатываемый файл
                        short_name_file = re.sub(r'[\[\]\'+()<> :"?*|\\/]', '_', name_file)
                        short_name_file = short_name_file[:23]
                        if short_name_file in set_used_named:
                            short_name_file = f'{short_name_file}_{count_sheets}'
                        else:
                            set_used_named.add(short_name_file)
                        wb_source = openpyxl.load_workbook(f'{dirpath}/{file}', data_only=False)  # открываем файл
                    except:
                        temp_error_df = pd.DataFrame(
                            data=[[f'{name_file}',
                                   f'Не удалось обработать файл. Возможно файл поврежден'
                                   ]],
                            columns=['Название файла',
                                     'Описание ошибки'])
                        error_df = pd.concat([error_df, temp_error_df], axis=0,
                                             ignore_index=True)
                        count_errors += 1
                        continue


                    for idx,sheet in enumerate(wb_source.sheetnames,1):
                        try:
                            name_sheet = str(idx)
                            target_sheet = main_wb.create_sheet(f'{short_name_file}_{name_sheet}') # создаем лист в сводном файле
                            source_sheet = wb_source[sheet]
                            #
                            copy_sheet(source_sheet, target_sheet)
                            count_sheets += 1
                        except:
                            temp_error_df = pd.DataFrame(
                                data=[[f'{name_file}',
                                       f'Не удалось обработать лист {sheet}. Возможно лист поврежден'
                                       ]],
                                columns=['Название файла',
                                         'Описание ошибки'])
                            error_df = pd.concat([error_df, temp_error_df], axis=0,
                                                 ignore_index=True)
                            count_errors += 1
                            continue

        if len(main_wb.sheetnames) == 1:
            main_wb.save(f'{end_folder}/Отсутствуют целые файлы {current_time}.xlsx')
        else:
            del main_wb['alone_sheet']
            main_wb.save(f'{end_folder}/Свод {current_time}.xlsx')
        if len(error_df) != 0:
            # если есть ошибки то сохраняем
            error_df.to_excel(f'{end_folder}/Ошибки {current_time}.xlsx')
    except NameError:
            messagebox.showerror('Веста Обработка таблиц и создание документов',
                                 f'Выберите папку с файлами и папку куда будет генерироваться результат')
    except FileNotFoundError:
        messagebox.showerror('Веста Обработка таблиц и создание документов',
                             f'Слишком длинный путь. Выберите в качестве конечной папку в корне диска или на рабочем столе')
    else:
        messagebox.showinfo('Веста Обработка таблиц и создание документов', f'Обработка завершена.')
















if __name__ == '__main__':
    main_data_folder = 'c:/Users/1/PycharmProjects/Elpida/data/Табличные отчеты Пример 1'
    main_end_folder = 'c:/Users/1/PycharmProjects/Elpida/data/СБОР результат'

    start_time = time.time()
    union_files_xlsx(main_data_folder, main_end_folder)
    end_time = time.time()
    elapsed_time = end_time - start_time
    print(f"Время выполнения: {elapsed_time:.6f} сек.")


    print('Lindy Booth')
